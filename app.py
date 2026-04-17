import os
from flask import Flask, render_template, request, jsonify, session, redirect, send_file
import database as db
import io
import datetime

app = Flask(__name__)
app.secret_key = 'collecte_dechets_secret_key_2024'

# Initialiser la base de données au démarrage
try:
    db.init_db()
except Exception as e:
    print(f"DB init error: {e}")


# --- Middleware d'authentification ---

@app.before_request
def check_auth():
    routes_publiques = ['login', 'static']
    if request.endpoint and request.endpoint not in routes_publiques and 'user' not in session:
        if request.is_json:
            return jsonify({'error': 'Non authentifié'}), 401
        return redirect('/login')


# --- Pages ---

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.get_json()
        user = db.verifier_login(data['login'], data['mot_de_passe'])
        if user:
            session['user'] = user['nom']
            session['user_id'] = user['id']
            return jsonify({'success': True})
        return jsonify({'error': 'Identifiants incorrects'}), 401
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')


@app.route('/')
def index():
    return render_template('index.html', user=session.get('user', ''))


# --- API Articles ---

@app.route('/api/articles', methods=['GET'])
def api_get_articles():
    return jsonify(db.get_articles())


@app.route('/api/articles', methods=['POST'])
def api_add_article():
    data = request.get_json()
    try:
        db.add_article(data['article'], data['unite'], float(data['prix']))
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/articles/<int:indice>', methods=['PUT'])
def api_update_article(indice):
    data = request.get_json()
    db.update_article(indice, data['article'], data['unite'], float(data['prix']))
    return jsonify({'success': True})


@app.route('/api/articles/<int:indice>', methods=['DELETE'])
def api_delete_article(indice):
    db.delete_article(indice)
    return jsonify({'success': True})


# --- API Sites Collectes ---

@app.route('/api/sites', methods=['GET'])
def api_get_sites():
    filtres = {
        'mois': request.args.get('mois', ''),
        'annee': request.args.get('annee', ''),
        'site': request.args.get('site', ''),
        'date_debut': request.args.get('date_debut', ''),
        'date_fin': request.args.get('date_fin', ''),
    }
    return jsonify(db.get_sites(filtres))


@app.route('/api/sites', methods=['POST'])
def api_add_site():
    data = request.get_json()
    num = db.add_site(
        data['date_collecte'], data['site'], data.get('contractant', ''),
        data.get('observation', ''), data.get('bon', ''),
        data.get('date_bon', ''), float(data.get('tonnage', 0))
    )
    # Ajouter les articles associés
    for item in data.get('articles', []):
        db.add_collecte(num, item['article'], float(item['qte']))
    return jsonify({'success': True, 'num': num})


@app.route('/api/sites/<int:num>', methods=['GET'])
def api_get_site(num):
    site = db.get_site_by_num(num)
    if not site:
        return jsonify({'error': 'Non trouvé'}), 404
    site['articles'] = db.get_collectes_by_site(num)
    return jsonify(site)


@app.route('/api/sites/<int:num>', methods=['PUT'])
def api_update_site(num):
    data = request.get_json()
    db.update_site(
        num, data['date_collecte'], data['site'], data.get('contractant', ''),
        data.get('observation', ''), data.get('bon', ''),
        data.get('date_bon', ''), float(data.get('tonnage', 0))
    )
    # Supprimer les anciens articles et réinsérer
    conn = db.get_db()
    conn.execute('DELETE FROM article_collecte WHERE num=?', (num,))
    conn.commit()
    conn.close()
    for item in data.get('articles', []):
        db.add_collecte(num, item['article'], float(item['qte']))
    return jsonify({'success': True})


@app.route('/api/sites/<int:num>', methods=['DELETE'])
def api_delete_site(num):
    db.delete_site(num)
    return jsonify({'success': True})


# --- API Analyse ---

@app.route('/api/stats/mensuelles')
def api_stats_mensuelles():
    annee = request.args.get('annee', '')
    return jsonify(db.get_stats_mensuelles(annee if annee else None))


@app.route('/api/stats/totaux')
def api_stats_totaux():
    return jsonify(db.get_totaux_par_article(
        request.args.get('date_debut'),
        request.args.get('date_fin')
    ))


@app.route('/api/recapitulatif')
def api_recapitulatif():
    date_debut = request.args.get('date_debut', '2000-01-01')
    date_fin = request.args.get('date_fin', '2099-12-31')
    return jsonify(db.get_recapitulatif(date_debut, date_fin))


# --- Export Excel ---

@app.route('/api/export/excel')
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    date_debut = request.args.get('date_debut', '2000-01-01')
    date_fin = request.args.get('date_fin', '2099-12-31')
    recap = db.get_recapitulatif(date_debut, date_fin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Récapitulatif"

    # Styles
    titre_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Titre
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Récapitulatif des collectes du {date_debut} au {date_fin}"
    ws['A1'].font = titre_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # En-têtes
    headers = ['Article', 'Unité', 'Prix unitaire', 'Quantité totale', 'Montant total', 'Nb collectes']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Données
    total_general = 0
    for i, r in enumerate(recap, 4):
        ws.cell(row=i, column=1, value=r['article']).border = border
        ws.cell(row=i, column=2, value=r['unite']).border = border
        ws.cell(row=i, column=3, value=r['prix']).border = border
        ws.cell(row=i, column=4, value=r['total_qte']).border = border
        ws.cell(row=i, column=5, value=r['total_montant']).border = border
        ws.cell(row=i, column=6, value=r['nb_collectes']).border = border
        total_general += r['total_montant'] or 0

    # Total
    row_total = len(recap) + 4
    ws.cell(row=row_total, column=4, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row_total, column=5, value=total_general).font = Font(bold=True)

    # Largeurs colonnes
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'recapitulatif_{date_debut}_{date_fin}.xlsx'
    )


# --- Export CSV ---

@app.route('/api/export/csv')
def export_csv():
    import csv

    date_debut = request.args.get('date_debut', '2000-01-01')
    date_fin = request.args.get('date_fin', '2099-12-31')

    sites = db.get_sites({'date_debut': date_debut, 'date_fin': date_fin})
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(['Num', 'Date', 'Site', 'Contractant', 'Article', 'Qté', 'Unité', 'Prix', 'Montant'])

    for site in sites:
        collectes = db.get_collectes_by_site(site['num'])
        for c in collectes:
            prix = c.get('prix') or 0
            writer.writerow([
                site['num'], site['date_collecte'], site['site'],
                site['contractant'], c['article'], c['qte'],
                c.get('unite', ''), prix, c['qte'] * prix
            ])

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'collectes_{date_debut}_{date_fin}.csv'
    )


# --- Export Facture PDF-like (HTML) ---

@app.route('/api/facture/<int:num>')
def facture(num):
    site = db.get_site_by_num(num)
    if not site:
        return "Non trouvé", 404
    articles = db.get_collectes_by_site(num)
    total = sum((a['qte'] * (a.get('prix') or 0)) for a in articles)
    return render_template('facture.html', site=site, articles=articles, total=total)


if __name__ == '__main__':
    print("=== Application Collecte de Déchets ===")
    print("Ouvrez http://localhost:5000 dans votre navigateur")
    print("Login par défaut : admin / admin")
    app.run(host='0.0.0.0', port=8080, debug=False)
